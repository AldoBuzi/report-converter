import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal, ROUND_HALF_UP
import re


from openpyxl.utils import get_column_letter

def autofit_columns(ws, dataframe, start_row):
    for i, col in enumerate(dataframe.columns, 1):
        column_letter = get_column_letter(i)
        # Get max length between header and all values
        max_length = len(str(col))

        for cell in ws[column_letter]:
            if cell.row < start_row:
                continue
            try:
                cell_value = str(cell.value)
                if cell_value:
                    max_length = max(max_length, len(cell_value))
            except:
                pass

        # Add a small margin for nicer spacing
        ws.column_dimensions[column_letter].width = max_length + 2

def autofit_rows(ws):
    for row in ws.iter_rows():
        if row[0].row < 4:
            continue
        max_lines = 1
        for cell in row:
            if cell.value:
                # Count newline characters to estimate required height
                lines = str(cell.value).count("\n") + 1
                max_lines = max(max_lines, lines)

        # Excel uses ~15 pt per line
        ws.row_dimensions[row[0].row].height = max_lines * 15


def normalize_val(val):
    if pd.isna(val) or val == "":
        return ""
    
    s_val = str(val).strip()
    
    try:
        d = Decimal(s_val)
    except:
        match = re.match(r"([-+]?\d*\.?\d+)", s_val)
        if match:
            d = Decimal(match.group(1))
        else:
            return s_val 
    
    rounded = d.quantize(Decimal('1'), rounding=ROUND_HALF_UP)
    return str(int(rounded))

def clean_currency(value):
    """
    Cleans currency strings like 'L 1,690.02' to float 1690.02.
    Handles already numeric values gracefully.
    """
    if isinstance(value, str):
        # Remove 'L', spaces, and commas
        clean_val = value.replace('L', '').replace(',', '').strip()
        try:
            return float(clean_val)
        except ValueError:
            return 0.0
    return value

def process_file(uploaded_file):
    # 1. Read the Input File
    # We skip the first 4 rows to get to the actual header (based on TransactionReport-3 structure)
    try:
        header_box_df = pd.read_excel(uploaded_file, nrows=4, header=None)
        search_criteria_right = header_box_df.iloc[1, 3] # Row 2, Column D (0-indexed)
        df = pd.read_excel(uploaded_file, skiprows=4)
    except Exception as e:
        st.error(f"Error reading CSV: {e}")
        return None, None

    # Define the columns we want to keep
    target_columns = [
        'Sl', 
        'Order Id', 
        'Restaurant', 
        'Customer Name', 
        'Delivery Charge', 
        #'Amount Received By',
        'Payment Status'
    ]

    # Check if columns exist
    missing_cols = [col for col in target_columns if col not in df.columns]
    if missing_cols:
        st.error(f"The uploaded file is missing these columns: {missing_cols}")
        st.write("Available columns:", list(df.columns))
        return None, None

    # Filter DataFrame
    df_filtered = df[target_columns].copy()

    # Convert "L 70.00" -> 70.00
    df_filtered['Delivery Charge'] = df_filtered['Delivery Charge'].apply(clean_currency).round(0)
    
    total_delivery =df_filtered['Delivery Charge'].apply(clean_currency).sum()

    # 5. formatting 'Delivery Charge' for output (optional, strictly numeric or keep currency?)
    # TransactionReport-4 shows strictly numeric for the column, so we keep it as float or simple string
    # But usually reports want clean numbers.

    return df_filtered, total_delivery, search_criteria_right

def generate_excel_bytes(df, total_val=None, search_criteria=""):
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        sheet_name = "Report"

        # --- Write DF starting later so header box fits ---
        start_row = 3
        df.to_excel(writer, sheet_name=sheet_name, startrow=start_row, index=False)

        workbook = writer.book
        ws = writer.sheets[sheet_name]

        # ============================
        # 1. BUILD TOP HEADER BOX
        # ============================
        table_width = df.shape[1]     # number of columns in table
        box_cols = 2                  # two-column layout

        # Merge cells for the left and right columns
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=table_width)
        ws.merge_cells(start_row=2, start_column=1,
                       end_row=2, end_column=3)
        ws.merge_cells(start_row=2, start_column=4,
                       end_row=2, end_column=table_width)
        ws.merge_cells(start_row=3, start_column=1,
                       end_row=3, end_column=3)
        ws.merge_cells(start_row=3, start_column=4,
                       end_row=3, end_column=table_width)
        # Insert content in the box
                # --- ROW 1: Title ---
        ws.cell(1, 1).value = "Order Transactions Report"
        ws.cell(1, 1).font = Font(bold=True, size=24)
        ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")

        # Set row height (50 pt)
        ws.row_dimensions[1].height = 50  

        # Set column A width to ~23 pt (Excel uses character width, so pt→width approx conversion)
        #ws.column_dimensions["A"].width = 23


        # --- ROW 2: Search Criteria ---
        ws.cell(2, 1).value = "Search Criteria"
        ws.cell(2, 1).font = Font(bold=True, size=11)
        ws.cell(2, 1).alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(2, 4).value = search_criteria
        ws.cell(2, 4).font = Font(bold=True, size=11)
        ws.cell(2, 4).alignment = Alignment(horizontal="left", vertical="center")

        # Row height
        ws.row_dimensions[2].height = 100

        # Column widths (converted from pt → Excel width approx)
        #ws.column_dimensions["A"].width = 100 / 7  # approx 100 pt
        #ws.column_dimensions["C"].width = 80 / 7   # approx 80 pt


        # --- ROW 3: Total label ---
        ws.cell(3, 1).value = "Total:"
        ws.cell(3, 1).font = Font(bold=True, size=11)
        ws.cell(3, 1).alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[3].height = 80
        #ws.column_dimensions["A"].width = 100 / 7
        
        # ===== Insert Excel SUM formula into header box =====
        amount_column = "Delivery Charge"  # change if needed
        amount_col_idx = df.columns.get_loc(amount_column) + 1
        amount_col_letter = get_column_letter(amount_col_idx)

        data_start = start_row + 2
        data_end = start_row + len(df) + 1

        sum_formula = f"=SUM({amount_col_letter}{data_start}:{amount_col_letter}{data_end + 1000000})"

        # Put formula in header box (right column)
        custom_format = '"L " #,##0.00'  # displays: L 71.40
        for r in range(data_start, data_end + 1):
                cell = ws.cell(r, amount_col_idx)
                # If value is None/NaN, leave blank
                if cell.value is None:
                    cell.value = None
                else:
                    # cell.value already numeric because df_copy is numeric
                    #cell.number_format = custom_format
                    cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(3, 4).value = sum_formula
        ws.cell(3, 4).number_format = custom_format
        ws.cell(3, 4).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(3, 4).font = Font(bold=True, size=11)
        
        # Add border around the whole box
        border_style = Side(border_style="thin", color="000000")
        box_border = Border(left=border_style, right=border_style,
                            top=border_style, bottom=border_style)

        for row in range(1, 4):
            for col in range(1, table_width + 1):
                ws.cell(row, col).border = box_border

        # ============================
        # 2. STYLE TABLE HEADER ROW
        # ============================
        header_row = start_row
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=header_row + 1, column=col_idx)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9D9D9", fill_type="solid")  # light gray

        # ============================
        # 3. CENTER ALL TABLE VALUES
        # ============================
        for r in range(start_row + 1, start_row + 1 + len(df) + 1):
            for c in range(1, table_width + 1):
                if(c == amount_col_idx):
                    continue
                ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center")

        # ============================
        # 4. AUTO COLUMN WIDTHS
        # ============================
        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            #ws.column_dimensions[col_letter].width = 15
            cell.alignment = Alignment(wrapText=True, vertical='top')
        
        autofit_rows(ws)
        autofit_columns(ws, df, 4)
    output.seek(0)
    return output.getvalue()

def generate_csv_string(df, total_val):
    """
    Manually constructs the CSV string to match the specific header format of Report-4.
    """
    output = io.StringIO()
    
    # Line 1: Report Title
    output.write("Order Transactions Report,,,,,,\n")
    
    # Line 2: Search Criteria (Placeholder or copied)
    # We can hardcode standard text or leave it blank
    output.write('Search Criteria,,,"Zone - Auto Generated\nFilter- Custom",,,\n')
    
    # Line 3: The Total Row
    # The format in Report-4 is: Total:,,,6141.52,,,
    # This aligns '6141.52' roughly under 'Delivery Charge' (5th column)
    # Columns: Sl(1), Order(2), Rest(3), Cust(4), Deliv(5)...
    # So we need 4 commas before the value.
    output.write(f"Total:,,,,{total_val:.2f},,,\n")
    
    # Line 4+: The Data Headers and Rows
    df.to_csv(output, index=False)
    
    return output.getvalue()

# --- STREAMLIT UI SETUP ---
st.set_page_config(page_title="Report Transformer", layout="centered")

st.title("📊 Transaction Report Converter")
st.markdown("""
Upload your **TransactionReport** file below. 
The app will filter columns, clean currency formats, calculate the total, and generate **TransactionReport-Converted**.
""")

# File Uploader (Drag and Drop)
uploaded_file = st.file_uploader("Drag and drop CSV file here", type=["xlsx", "csv"])

if uploaded_file is not None:
    st.success("File uploaded successfully!")
    
    # Process
    df_result, total_val, search_criteria = process_file(uploaded_file)
    
    if df_result is not None:
        # Show Preview
        st.subheader("Preview of Generated Report")
        st.metric(label="Calculated Total Delivery Charge", value=f"{total_val:,.2f}")
        st.dataframe(df_result.head(10))
        
        # Generate CSV String
        csv_data = generate_excel_bytes(df_result, total_val, search_criteria)
        
        # Download Button
        st.download_button(
            label="📥 Download TransactionReport",
            data=csv_data,
            file_name="TransactionReport_Converted.xlsx",
            mime="text/csv"
        )